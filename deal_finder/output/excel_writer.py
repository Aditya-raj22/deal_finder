"""Excel output writer."""

from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font

from ..models import Deal, ExcelRow


class ExcelWriter:
    """Write deals to Excel file with required schema."""

    # Column headers (exact order from spec)
    HEADERS = [
        "Date Announced",
        "Target / Partner",
        "Acquirer / Partner",
        "Upfront Value (M USD)",
        "Contingent Payment (M USD)",
        "Total Deal Value (M USD)",
        "Upfront as % of Total Value",
        "Phase of Lead Asset at Announcement",
        "Therapeutic Area",
        "Secondary Areas",
        "Asset / Focus",
        "Deal Type (M&A or Partnership)",
        "Geography of Target",
        "Source URL",
        "Needs Review",
    ]

    def __init__(self):
        pass

    def write(self, deals: List[Deal], output_path: str) -> None:
        """Write deals to Excel file."""
        # Convert deals to Excel rows
        rows = [ExcelRow.from_deal(deal) for deal in deals]

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Deals"

        # Write headers
        ws.append(self.HEADERS)

        # Make headers bold
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Write data rows
        for row in rows:
            ws.append(
                [
                    row.date_announced,
                    row.target,
                    row.acquirer,
                    float(row.upfront_value_m_usd) if row.upfront_value_m_usd else None,
                    (
                        float(row.contingent_payment_m_usd)
                        if row.contingent_payment_m_usd
                        else None
                    ),
                    (
                        float(row.total_deal_value_m_usd)
                        if row.total_deal_value_m_usd
                        else None
                    ),
                    float(row.upfront_as_pct_total) if row.upfront_as_pct_total else None,
                    row.phase_at_announcement,
                    row.therapeutic_area,
                    row.secondary_areas,
                    row.asset_focus,
                    row.deal_type,
                    row.geography,
                    row.source_url,
                    "TRUE" if row.needs_review else "FALSE",
                ]
            )

        # Format date column
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
            for cell in row:
                if cell.value:
                    cell.number_format = "YYYY-MM-DD"

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save workbook
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
