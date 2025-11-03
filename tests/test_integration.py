"""Integration tests."""

import os
import tempfile
from pathlib import Path

import pytest
from freezegun import freeze_time

from deal_finder.config_loader import Config
from deal_finder.output.excel_writer import ExcelWriter
from deal_finder.models import Deal, DealTypeDetailed, DevelopmentStage
from datetime import date


@pytest.fixture
def temp_output_dir():
    with tempfile.TemporaryDirectory() as tmpdir:
        yield tmpdir


@pytest.fixture
def sample_deals():
    return [
        Deal(
            date_announced=date(2023, 6, 15),
            target="BioCompany A",
            acquirer="Pharma Corp",
            stage=DevelopmentStage.PRECLINICAL,
            therapeutic_area="immunology_inflammation",
            asset_focus="IL-6 inhibitor",
            deal_type_detailed=DealTypeDetailed.MA,
            source_url="https://example.com/deal1",
            needs_review=False,
            timestamp_utc="2023-06-15T00:00:00Z",
        ),
        Deal(
            date_announced=date(2023, 7, 20),
            target="BioCompany B",
            acquirer="BioCompany C",
            stage=DevelopmentStage.PHASE_1,
            therapeutic_area="immunology_inflammation",
            asset_focus="TNF-alpha blocker",
            deal_type_detailed=DealTypeDetailed.PARTNERSHIP,
            source_url="https://example.com/deal2",
            needs_review=True,
            timestamp_utc="2023-07-20T00:00:00Z",
        ),
    ]


def test_excel_writer_creates_file(temp_output_dir, sample_deals):
    writer = ExcelWriter()
    output_path = os.path.join(temp_output_dir, "test_output.xlsx")

    writer.write(sample_deals, output_path)

    assert Path(output_path).exists()


def test_excel_writer_correct_columns(temp_output_dir, sample_deals):
    from openpyxl import load_workbook

    writer = ExcelWriter()
    output_path = os.path.join(temp_output_dir, "test_output.xlsx")

    writer.write(sample_deals, output_path)

    # Load and check headers
    wb = load_workbook(output_path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]

    assert headers == ExcelWriter.HEADERS


def test_excel_writer_correct_row_count(temp_output_dir, sample_deals):
    from openpyxl import load_workbook

    writer = ExcelWriter()
    output_path = os.path.join(temp_output_dir, "test_output.xlsx")

    writer.write(sample_deals, output_path)

    wb = load_workbook(output_path)
    ws = wb.active

    # Headers + 2 data rows
    assert ws.max_row == 3


def test_excel_writer_boolean_format(temp_output_dir, sample_deals):
    from openpyxl import load_workbook

    writer = ExcelWriter()
    output_path = os.path.join(temp_output_dir, "test_output.xlsx")

    writer.write(sample_deals, output_path)

    wb = load_workbook(output_path)
    ws = wb.active

    # Check Needs Review column (last column)
    needs_review_col_idx = len(ExcelWriter.HEADERS)

    # First deal: needs_review=False
    assert ws.cell(row=2, column=needs_review_col_idx).value == "FALSE"

    # Second deal: needs_review=True
    assert ws.cell(row=3, column=needs_review_col_idx).value == "TRUE"


@freeze_time("2023-08-01")
def test_config_end_date_resolution():
    config = Config(THERAPEUTIC_AREA="immunology_inflammation")

    assert config.end_date_resolved == "2023-08-01"


def test_idempotency_same_input_same_output(temp_output_dir, sample_deals):
    """Test that running with same inputs produces identical output."""
    writer = ExcelWriter()
    output_path1 = os.path.join(temp_output_dir, "output1.xlsx")
    output_path2 = os.path.join(temp_output_dir, "output2.xlsx")

    # Write twice
    writer.write(sample_deals, output_path1)
    writer.write(sample_deals, output_path2)

    # Read both files and compare
    from openpyxl import load_workbook

    wb1 = load_workbook(output_path1)
    wb2 = load_workbook(output_path2)

    ws1 = wb1.active
    ws2 = wb2.active

    # Compare all cells
    for row1, row2 in zip(ws1.iter_rows(), ws2.iter_rows()):
        for cell1, cell2 in zip(row1, row2):
            assert cell1.value == cell2.value
